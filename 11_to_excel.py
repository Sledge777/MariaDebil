import pandas as pd
from sqlalchemy import create_engine
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# Параметры подключения
host = "localhost"
port = 3306
user = "root"
password = "Ms_phone_fox62052!"
database = "ms_phone"

# Подключение к базе
engine = create_engine(f"mysql+pymysql://{user}:{password}@{host}:{port}/{database}")

# Получаем список всех таблиц в базе
with engine.connect() as conn:
    tables_df = pd.read_sql("SHOW TABLES", conn)
    table_names = tables_df.iloc[:, 0].tolist()

# Имя выходного Excel-файла
output_file = f"{database}_dump.xlsx"

# Выгрузка всех таблиц
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for table_name in table_names:
        df = pd.read_sql(f"SELECT * FROM `{table_name}`", engine)
        df.to_excel(writer, sheet_name=table_name[:31], index=False)

# Стилизация: заголовки оранжевые, автоширина столбцов
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Оранжевая заливка заголовков
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # Автоширина
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# Сохраняем изменения
wb.save(output_file)

print(f"Готово! Файл {output_file} создан с оформлением.")
